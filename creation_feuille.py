from openpyxl import Workbook
classeur=Workbook()
etudiants=classeur.create_sheet("etudiants")
cours=classeur.create_sheet("cours")
notes=classeur.create_sheet("notes")
etudiants.append(["Matricule","Nom","Postnom","Prenom","DateNaissance","LieuNaissance"])
cours.append(["Sigle","IntituleCours"])
notes.append(["Matricule","Sigle","Note"])
classeur.save("E:/Projet openpy/exemple.xlsx")
