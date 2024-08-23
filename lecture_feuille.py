from openpyxl import load_workbook
classeur=load_workbook("exemple.xlsx")
etudiants=classeur["etudiants"]
matri=4422
for i,row in enumerate(etudiants.rows):
    val=row
    for k in val:
        print(val.value)
    matricule=etudiants.cell(row=i+1,column=1).value
    if(matricule==matri):
        print("Le matricule existe déjà")

